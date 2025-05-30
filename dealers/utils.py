import os
import openpyxl
from openpyxl.styles import Font, Alignment
from django.conf import settings
from django.utils import timezone
from .models import Dealer, DealerStockNorm, Part
from datetime import datetime
from io import BytesIO
import zipfile


def generate_dealer_files():
    """Генерирует Excel-файлы для всех активных дилеров"""
    base_dir = os.path.join(settings.MEDIA_ROOT, 'dealer_files')
    os.makedirs(base_dir, exist_ok=True)

    dealers = Dealer.objects.filter(is_active=True)
    generated_files = []

    for dealer in dealers:
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Остатки"

        # Заголовки
        headers = ['Артикул', 'Наименование товара', 'Дилерская цена', 'Текущее количество на складе']
        ws.append(headers)

        # Форматирование заголовков
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Получаем все нормы для дилера
        norms = DealerStockNorm.objects.filter(dealer=dealer).select_related('part')

        for norm in norms:
            row = [
                norm.part.original_number,
                norm.part.name,
                "",  # Цена будет заполняться вручную
                norm.current_stock  # Текущий запас
            ]
            ws.append(row)

        # Сохраняем файл
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"dealer_{dealer.id}_{timestamp}.xlsx"
        file_path = os.path.join(base_dir, file_name)
        wb.save(file_path)

        generated_files.append(file_path)

    return generated_files


def process_dealer_files():
    """Обрабатывает файлы, загруженные дилерами"""
    upload_dir = os.path.join(settings.MEDIA_ROOT, 'dealer_uploads')
    processed_dir = os.path.join(settings.MEDIA_ROOT, 'dealer_uploads', 'processed')
    error_dir = os.path.join(settings.MEDIA_ROOT, 'dealer_uploads', 'errors')

    os.makedirs(upload_dir, exist_ok=True)
    os.makedirs(processed_dir, exist_ok=True)
    os.makedirs(error_dir, exist_ok=True)

    processed_count = 0
    for file_name in os.listdir(upload_dir):
        if not file_name.endswith('.xlsx'):
            continue

        file_path = os.path.join(upload_dir, file_name)

        try:
            # Извлекаем ID дилера из имени файла
            parts = file_name.split('_')
            if len(parts) < 2 or not parts[1].isdigit():
                raise ValueError(f"Неверный формат имени файла: {file_name}")

            dealer_id = int(parts[1])
            dealer = Dealer.objects.get(id=dealer_id)

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            updated_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue

                part_number = str(row[0]).strip()
                try:
                    current_stock = int(row[3]) if row[3] is not None else 0
                except (TypeError, ValueError):
                    current_stock = 0

                try:
                    part = Part.objects.get(original_number=part_number)
                    norm, created = DealerStockNorm.objects.get_or_create(
                        dealer=dealer,
                        part=part,
                        defaults={'norm': 0, 'current_stock': current_stock}
                    )

                    if not created:
                        norm.current_stock = current_stock
                        norm.save()

                    updated_count += 1
                except Part.DoesNotExist:
                    continue

            # Перемещаем обработанный файл
            os.rename(file_path, os.path.join(processed_dir, file_name))
            processed_count += 1

        except Exception as e:
            # Перемещаем файл с ошибкой
            os.rename(file_path, os.path.join(error_dir, file_name))

    return processed_count


def generate_dealer_report_excel(report_id, include_inactive=False):
    """Генерирует Excel-отчет по распределению товаров"""
    from .models import DealerDistributionReport

    report = DealerDistributionReport.objects.get(id=report_id)

    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Распределение"

    # Заголовки
    dealers = Dealer.objects.all() if include_inactive else Dealer.objects.filter(is_active=True)
    headers = ['Артикул', 'Наименование товара', 'Общий остаток на складе']
    for dealer in dealers:
        headers.extend([f"{dealer.customer.name} (норма)", f"{dealer.customer.name} (текущий запас)",
                        f"{dealer.customer.name} (спрос)", f"{dealer.customer.name} (к отправке)"])

    ws.append(headers)

    # Форматирование заголовков
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Получаем все детали, для которых есть нормы
    part_ids = DealerStockNorm.objects.values_list('part_id', flat=True).distinct()
    parts = Part.objects.filter(id__in=part_ids)

    # Заполняем данные
    for part in parts:
        # Получаем общий остаток на складе (заглушка - в реальности из системы)
        total_stock = 1000  # Здесь должна быть реальная логика

        row = [
            part.original_number,
            part.name,
            total_stock
        ]

        for dealer in dealers:
            try:
                norm = DealerStockNorm.objects.get(part=part, dealer=dealer)
                demand = max(0, norm.norm - norm.current_stock)
                quantity_to_send = min(demand, total_stock)
                total_stock -= quantity_to_send

                row.extend([
                    norm.norm,
                    norm.current_stock,
                    demand,
                    quantity_to_send
                ])
            except DealerStockNorm.DoesNotExist:
                row.extend([0, 0, 0, 0])

        ws.append(row)

    # Сохраняем файл
    file_name = f"dealer_distribution_{report_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, 'dealer_reports', file_name)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)

    # Обновляем отчет
    report.report_file.name = os.path.join('dealer_reports', file_name)
    report.save()

    return file_path


def generate_dealer_waybill_pdf(report_id, dealer_id):
    """Генерирует PDF транспортной накладной для дилера"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from .models import DealerDistributionReport, Dealer, Part

    report = DealerDistributionReport.objects.get(id=report_id)
    dealer = Dealer.objects.get(id=dealer_id)

    # Создаем буфер для PDF
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)

    # Настройки стилей
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.white,
        alignment=TA_CENTER
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )
    bold_style = ParagraphStyle(
        'Bold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )

    # Заголовок
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(A4[0] / 2, A4[1] - 2 * cm, "ТРАНСПОРТНАЯ НАКЛАДНАЯ")
    c.setFont("Helvetica", 10)
    c.drawCentredString(A4[0] / 2, A4[1] - 2.5 * cm,
                        f"№ {report.id}/{dealer.id} от {report.created_at.strftime('%d.%m.%Y')}")

    # Информация о компании
    y_position = A4[1] - 4 * cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y_position, "Грузоотправитель:")
    c.setFont("Helvetica", 10)
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, "ООО 'Автозапчасти'")
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, "Юридический адрес: г. Москва, ул. Автозапчастей, д. 1")
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, "ИНН 1234567890, КПП 123456789")
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, "Тел.: +7 (495) 123-45-67")

    # Информация о дилере
    y_position -= 0.7 * cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2 * cm, y_position, "Грузополучатель:")
    c.setFont("Helvetica", 10)
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, dealer.customer.name)
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, f"Контактное лицо: {dealer.contact_person}")
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, f"Тел.: {dealer.customer.phone}")
    y_position -= 0.5 * cm
    c.drawString(3 * cm, y_position, f"Email: {dealer.email}")

    # Таблица с товарами
    y_position -= 1 * cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, y_position, "Перечень товаров:")
    y_position -= 0.5 * cm

    # Заглушка для данных - в реальном приложении данные должны браться из отчета
    items = [
        {"part": Part(original_number="ABC123", name="Масляный фильтр"), "quantity": 5, "price": 1500.00},
        {"part": Part(original_number="DEF456", name="Воздушный фильтр"), "quantity": 3, "price": 800.00},
        {"part": Part(original_number="GHI789", name="Тормозные колодки"), "quantity": 2, "price": 3200.00},
    ]

    data = []
    headers = [
        "№",
        "Артикул",
        "Наименование товара",
        "Кол-во",
        "Цена за ед., руб.",
        "Сумма, руб."
    ]
    data.append([Paragraph(header, header_style) for header in headers])

    total = 0
    for i, item in enumerate(items, 1):
        item_sum = item['quantity'] * item['price']
        total += item_sum

        row = [
            Paragraph(str(i), cell_style),
            Paragraph(item['part'].original_number, cell_style),
            Paragraph(item['part'].name, cell_style),
            Paragraph(str(item['quantity']), cell_style),
            Paragraph(f"{item['price']:.2f}", cell_style),
            Paragraph(f"{item_sum:.2f}", cell_style),
        ]
        data.append(row)

    # Итоговая строка
    data.append([
        "", "", "", "",
        Paragraph("Итого:", bold_style),
        Paragraph(f"{total:.2f}", bold_style)
    ])

    # Создаем таблицу
    col_widths = [1 * cm, 2.5 * cm, 7 * cm, 2 * cm, 3 * cm, 3 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Стили таблицы
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9D9D9')),
    ])
    table.setStyle(table_style)

    # Размещаем таблицу
    table.wrapOn(c, A4[0] - 4 * cm, A4[1] - 10 * cm)
    table.drawOn(c, 2 * cm, y_position - 1 * cm - table._height)

    # Подписи
    y_bottom = 2 * cm
    c.setFont("Helvetica", 10)
    c.drawString(2 * cm, y_bottom + 1.5 * cm, "Отпуск разрешил: ___________________ / Петров И.С. /")
    c.drawString(2 * cm, y_bottom + 1 * cm, "должность, подпись, ФИО")

    c.drawString(10 * cm, y_bottom + 1.5 * cm, "Груз принял: ___________________ /")
    c.drawString(10 * cm, y_bottom + 1 * cm, "должность, подпись, ФИО")

    c.drawString(2 * cm, y_bottom, f"Всего отпущено на сумму: {total:.2f} руб.")
    c.drawString(10 * cm, y_bottom, "М.П.")

    c.showPage()
    c.save()

    buffer.seek(0)
    return buffer