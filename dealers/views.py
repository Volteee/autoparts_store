from django import forms
from django.views.generic import TemplateView
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.http import FileResponse
from .forms import DealerForm, DealerStockNormForm, FileUploadForm, DealerWaybillGenerationForm
from .utils import generate_dealer_files, process_dealer_files, generate_dealer_report_excel, \
    generate_dealer_waybill_pdf
from django.conf import settings
from django.urls import reverse
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import redirect, render
from .models import Dealer, DealerStockNorm, DealerDistributionReport, Part, DealerWaybill
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from core.mixins import RoleRequiredMixin
import os


class DealerDistributionView(TemplateView):
    template_name = 'dealers/dealer_distribution.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получаем всех активных дилеров
        dealers = Dealer.objects.filter(is_active=True).prefetch_related('stock_norms')

        # Получаем все детали, для которых есть нормы у дилеров
        part_ids = DealerStockNorm.objects.values_list('part_id', flat=True).distinct()
        parts = Part.objects.filter(id__in=part_ids).prefetch_related('dealer_norms')

        # Собираем данные для отчета
        report_data = []
        for part in parts:
            part_data = {
                'part': part,
                'dealers': [],
                'total_stock': 100  # Заглушка: реальный остаток на главном складе
            }

            for dealer in dealers:
                try:
                    norm = dealer.stock_norms.get(part=part)
                    current_stock = norm.current_stock if hasattr(norm, 'current_stock') else 0
                    demand = max(0, norm.norm - current_stock)
                    quantity_to_send = min(demand, part_data['total_stock'])
                    part_data['total_stock'] -= quantity_to_send

                    part_data['dealers'].append({
                        'dealer': dealer,
                        'norm': norm,
                        'demand': demand,
                        'quantity_to_send': quantity_to_send
                    })
                except DealerStockNorm.DoesNotExist:
                    part_data['dealers'].append({
                        'dealer': dealer,
                        'norm': None,
                        'demand': 0,
                        'quantity_to_send': 0
                    })

            report_data.append(part_data)

        context['report_data'] = report_data
        context['dealers'] = dealers
        return context

    def post(self, request, *args, **kwargs):
        # Сохраняем отчет в базе данных
        report = DealerDistributionReport.objects.create()

        # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Распределение"

        # Заголовки
        headers = ['Артикул', 'Наименование товара', 'Общий остаток']
        dealers = Dealer.objects.filter(is_active=True)
        for dealer in dealers:
            headers.extend([f"{dealer.customer.name} (спрос)", f"{dealer.customer.name} (отправка)"])

        ws.append(headers)

        # Стили для заголовков
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        # Заполняем данными
        context = self.get_context_data()
        for item in context['report_data']:
            row = [
                item['part'].original_number,
                item['part'].name,
                item['total_stock']
            ]

            for dealer_data in item['dealers']:
                row.append(dealer_data['demand'])
                row.append(dealer_data['quantity_to_send'])

            ws.append(row)

        # Сохраняем файл
        file_name = f"dealer_distribution_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join('media', 'dealer_reports', file_name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)

        # Обновляем отчет
        report.report_file = file_path
        report.save()

        messages.success(request, 'Отчет успешно сгенерирован!')
        return redirect('dealers/reports/')


# Дилеры
class DealerListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = Dealer
    template_name = 'dealers/dealer_list.html'
    context_object_name = 'dealers'
    allowed_roles = ['supply_manager']


class DealerCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    model = Dealer
    form_class = DealerForm
    template_name = 'dealers/dealer_form.html'
    success_url = reverse_lazy('dealer_list')
    allowed_roles = ['supply_manager']

    # Добавляем обработку формы
    def form_valid(self, form):
        # Устанавливаем связь с клиентом
        form.instance.customer = form.cleaned_data['customer']
        return super().form_valid(form)

class DealerUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    model = Dealer
    form_class = DealerForm
    template_name = 'dealers/dealer_form.html'
    success_url = reverse_lazy('dealer_list')
    allowed_roles = ['supply_manager']

    # Добавляем обработку формы
    def form_valid(self, form):
        form.instance.customer = form.cleaned_data['customer']
        return super().form_valid(form)


class DealerDeleteView(LoginRequiredMixin, RoleRequiredMixin, DeleteView):
    model = Dealer
    template_name = 'dealers/dealer_confirm_delete.html'
    success_url = reverse_lazy('dealer_list')
    allowed_roles = ['supply_manager']


# Нормы запасов
class DealerStockNormListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = DealerStockNorm
    template_name = 'dealers/stock_norm_list.html'
    context_object_name = 'norms'
    allowed_roles = ['supply_manager']


class DealerStockNormCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    model = DealerStockNorm
    form_class = DealerStockNormForm
    template_name = 'dealers/stock_norm_form.html'
    success_url = reverse_lazy('stock_norm_list')
    allowed_roles = ['supply_manager']


class DealerStockNormUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    model = DealerStockNorm
    form_class = DealerStockNormForm
    template_name = 'dealers/stock_norm_form.html'
    success_url = reverse_lazy('stock_norm_list')
    allowed_roles = ['supply_manager']


# Работа с файлами
class GenerateDealerFilesView(LoginRequiredMixin, RoleRequiredMixin, FormView):
    template_name = 'dealers/generate_files.html'
    form_class = forms.Form
    success_url = reverse_lazy('dealer_list')
    allowed_roles = ['supply_manager']

    def form_valid(self, form):
        generate_dealer_files()
        return super().form_valid(form)


class ProcessDealerFilesView(LoginRequiredMixin, RoleRequiredMixin, FormView):
    template_name = 'dealers/process_files.html'
    form_class = forms.Form
    success_url = reverse_lazy('dealer_list')
    allowed_roles = ['supply_manager']

    def form_valid(self, form):
        count = process_dealer_files()
        return super().form_valid(form)


class FileUploadView(LoginRequiredMixin, RoleRequiredMixin, FormView):
    template_name = 'dealers/upload_file.html'
    form_class = FileUploadForm
    success_url = reverse_lazy('dealer_list')
    allowed_roles = ['supply_manager']

    def form_valid(self, form):
        file = form.cleaned_data['dealer_file']
        # Сохраняем файл для обработки
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'dealer_uploads')
        os.makedirs(upload_dir, exist_ok=True)
        with open(os.path.join(upload_dir, file.name), 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        return super().form_valid(form)


# Отчеты
class DealerDistributionReportListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = DealerDistributionReport
    template_name = 'dealers/report_list.html'
    context_object_name = 'reports'
    allowed_roles = ['supply_manager']

    def get_queryset(self):
        return DealerDistributionReport.objects.order_by('-created_at')


class GenerateReportView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    model = DealerDistributionReport
    template_name = 'dealers/generate_report.html'
    allowed_roles = ['supply_manager']

    def get(self, request, *args, **kwargs):
        report = self.get_object()
        file_path = generate_dealer_report_excel(report.id)
        return FileResponse(open(file_path, 'rb'), as_attachment=True)


# Накладные
class DealerWaybillListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = DealerWaybill
    template_name = 'dealers/waybill_list.html'
    context_object_name = 'waybills'
    allowed_roles = ['supply_manager']

    def get_queryset(self):
        return DealerWaybill.objects.order_by('-created_at')


class GenerateWaybillView(LoginRequiredMixin, RoleRequiredMixin, FormView):
    template_name = 'dealers/generate_waybill.html'
    form_class = DealerWaybillGenerationForm
    allowed_roles = ['supply_manager']

    def form_valid(self, form):
        report = form.cleaned_data['report']
        dealer = form.cleaned_data['dealer']
        pdf_buffer = generate_dealer_waybill_pdf(report.id, dealer.id)
        return FileResponse(pdf_buffer, as_attachment=True, filename=f'waybill_{dealer.id}.pdf')